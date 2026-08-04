"""ExcelDataLoader - Excel 数据加载器

使用 calamine (Rust) 引擎快速读取，fallback 到 openpyxl iter_rows。
"""

from __future__ import annotations
import os
import gc
from typing import Callable
import numpy as np
import pandas as pd
from src.data.base_loader import BaseDataLoader
from src.core.config import _UNIT_KEYWORDS, UNIT_KEYWORD_RATIO_THRESHOLD, EXCEL_MAX_SCAN_ROWS
from src.core.logger import get_logger

logger = get_logger("data.excel_loader")


class ExcelDataLoader(BaseDataLoader):
    """Excel 数据加载器，继承自 BaseDataLoader。

    优先使用 calamine (Rust) 引擎读取，不可用时回退到优化后的 openpyxl。
    """

    LOADER_TYPE = "excel"

    # 分块大小：openpyxl 流式读取时的批大小
    CHUNK_SIZE = 10000

    def __init__(
        self,
        file_path: str,
        *,
        sheet_name: str | int = 0,
        downcast_float: bool = True,
        desc_rows: int | None = None,
        has_unit: bool | None = None,
        _progress: Callable | None = None,
    ):
        """初始化 Excel 数据加载器

        Args:
            file_path: Excel 文件路径
            sheet_name: Sheet 名称或索引（0-based）
            downcast_float: 是否下转换浮点数类型
            desc_rows: 描述行数量，None 表示自动检测
            has_unit: 是否包含单位行，None 表示自动检测
            _progress: 进度回调函数
        """
        super().__init__()

        self._path = file_path
        self.file_size = os.path.getsize(file_path)
        self._sheet_name = sheet_name
        self.downcast_float = downcast_float
        self._progress_cb = _progress

        logger.info("开始加载 Excel: %s (%.1f MB)", file_path, self.file_size / 1024 / 1024)

        import openpyxl
        self._wb = openpyxl.load_workbook(
            file_path, read_only=True, data_only=True
        )

        try:
            # 解析 sheet_name（支持字符串名或索引）
            if isinstance(sheet_name, int):
                ws_name = self._wb.sheetnames[sheet_name]
            else:
                ws_name = sheet_name
            self._ws = self._wb[ws_name]

            if self._progress_cb:
                self._progress_cb(5)

            # 自动检测 desc_rows
            if desc_rows is None:
                try:
                    self.desc_rows, has_unit = self._auto_detect_format(has_unit)
                    logger.info(
                        "Excel auto_detect: desc_rows=%d, has_unit=%s (sheet=%s)",
                        self.desc_rows, has_unit, ws_name,
                    )
                except Exception as e:
                    logger.warning("Excel 自动检测失败 (%s)，使用默认值 desc_rows=0", e)
                    self.desc_rows = 0
            else:
                self.desc_rows = desc_rows

            # 读取表头 + 自动检测 has_unit
            self._var_names, self._units, self.has_unit = self._load_header_units(has_unit)

            if self._progress_cb:
                self._progress_cb(10)

            # 读取数据：优先使用 calamine (Rust)，fallback 到优化后的 openpyxl
            self._df = self._read_data()

            # 时间列推断
            self._infer_time_columns()

            # 后处理：合并 downcast + validity 为单次遍历
            self._df_validity = self._postprocess_columns(
                downcast=downcast_float
            )
        finally:
            # 关闭 workbook 释放资源（finally 保证异常路径也关闭，避免文件句柄泄漏）
            self._wb.close()

        if self._progress_cb:
            self._progress_cb(100)

        logger.info(
            "Excel 加载完成: %s (sheet=%s, %d 行, %d 列)",
            file_path, ws_name, len(self._df), len(self._var_names),
        )

    def _load_header_units(self, has_unit_hint: bool | None):
        """加载表头和单位信息，支持 has_unit 自动检测"""
        header_row_idx = self.desc_rows + 1  # openpyxl 1-based 行号
        rows = []
        for row in self._ws.iter_rows(
            min_row=header_row_idx, max_row=header_row_idx + 1, values_only=True
        ):
            rows.append(row)

        if not rows:
            raise ValueError("Excel 文件标题行为空")

        # 解析列名（清理前后空格和换行符）
        var_names = [
            str(cell).strip().replace("\n", " ").replace("\r", "")
            if cell is not None else f"Column_{i}"
            for i, cell in enumerate(rows[0])
        ]
        var_names = self._make_unique(var_names)

        # has_unit 自动检测
        if has_unit_hint is not None:
            actual_has_unit = has_unit_hint
        elif len(rows) >= 2:
            unit_row = rows[1]
            unit_hit = 0
            total = 0
            for cell in unit_row:
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                if not cell_str:
                    continue
                total += 1
                cell_lower = cell_str.lower()
                for keyword in _UNIT_KEYWORDS:
                    if keyword.lower() in cell_lower:
                        unit_hit += 1
                        break
            actual_has_unit = (
                total > 0 and (unit_hit / total) > UNIT_KEYWORD_RATIO_THRESHOLD
            )
        else:
            actual_has_unit = False

        if actual_has_unit and len(rows) >= 2:
            units = {
                name: (str(cell).strip() if cell is not None else "-")
                for name, cell in zip(var_names, rows[1])
            }
        else:
            units = {name: "-" for name in var_names}

        return var_names, units, actual_has_unit

    # 列数跳跃检测阈值：当候选标题行的列数超过此前候选列数的 N 倍时，
    # 判定前面的行为元数据描述区，当前行为真正的标题行。
    _COLUMN_JUMP_FACTOR = 3

    @staticmethod
    def _detect_header_row_from_rows(rows: list[tuple]) -> int:
        """从 Excel 行数据中定位标题行（列数跳跃检测策略）

        Excel 有严格的列网格结构，这与 CSV 有本质区别：
        - CSV 的元数据行可能占满整行宽度（以分隔符数量衡量）
        - Excel 的元数据行（如 UniPlot key=value）通常仅填充前 1-2 列，
          而真正的标题行会填满所有数据列。

        因此不采用 CSV 的「首命中即返回」，而是追踪候选行列数：
        当后续候选行的列数发生数量级跳跃（≥ COLUMN_JUMP_FACTOR 倍），
        说明前面的候选在元数据区，跳跃点才是真正标题行。
        若无跳跃，则返回唯一（或列数最多）的候选行。

        Args:
            rows: openpyxl iter_rows(values_only=True) 读取的前 N 行

        Returns:
            标题行在 rows 中的 0-based 索引，未找到时返回 0
        """
        candidate_idx = -1
        candidate_cols = 0

        for idx, row in enumerate(rows):
            str_count = 0
            numeric_count = 0
            non_null = 0

            for cell in row:
                if cell is None:
                    continue
                non_null += 1
                if isinstance(cell, str):
                    cell_str = cell.strip()
                    if not cell_str:
                        non_null -= 1  # 纯空白不计入有效列
                        continue
                    try:
                        float(cell_str)
                        numeric_count += 1
                    except ValueError:
                        str_count += 1
                elif isinstance(cell, (int, float)):
                    numeric_count += 1
                else:
                    str_count += 1

            total_valid = str_count + numeric_count
            if total_valid < 2:
                continue
            if str_count / total_valid <= 0.5:
                continue

            # 找到第一个候选
            if candidate_idx == -1:
                candidate_idx = idx
                candidate_cols = non_null
                continue

            # 列数发生数量级跳跃 → 前面的候选在元数据区，当前行才是标题
            if non_null >= candidate_cols * ExcelDataLoader._COLUMN_JUMP_FACTOR:
                return idx

            # 列数更多但不足够跳跃 → 更新为更优候选
            if non_null > candidate_cols:
                candidate_idx = idx
                candidate_cols = non_null

        return candidate_idx if candidate_idx != -1 else 0

    def _auto_detect_format(self, has_unit_hint: bool | None) -> tuple[int, bool | None]:
        """自动检测 Excel 文件的描述行数（实例方法，复用已打开的 self._ws）

        扫描前 EXCEL_MAX_SCAN_ROWS 行，用 _detect_header_row_from_rows
        定位标题行位置。has_unit 不做独立检测，保持 has_unit_hint
        原值由后续 _load_header_units 处理。

        Args:
            has_unit_hint: 外部传入的 has_unit 提示

        Returns:
            (desc_rows, has_unit_hint)
        """
        rows = list(
            self._ws.iter_rows(
                max_row=EXCEL_MAX_SCAN_ROWS, values_only=True
            )
        )
        header_idx = self._detect_header_row_from_rows(rows)
        return header_idx, has_unit_hint

    def _read_data(self) -> pd.DataFrame:
        """读取数据：优先 calamine (Rust)，不可用时回退到优化后的 openpyxl"""
        try:
            return self._read_with_calamine()
        except ImportError:
            logger.info("calamine 不可用，使用优化后的 openpyxl 读取")
        except Exception as e:
            logger.warning("calamine 读取失败 (%s)，fallback 到 openpyxl", e)

        return self._read_chunks()

    def _read_with_calamine(self) -> pd.DataFrame:
        """使用 calamine (Rust) 引擎快速读取 Excel 数据。

        calamine 读取缓存值（与 openpyxl data_only=True 行为一致），
        不支持公式重新求值，但对正常保存的 Excel 文件无影响。
        """
        data_start = self.desc_rows + (3 if self.has_unit else 2)  # 1-based

        df = pd.read_excel(
            self._path,
            sheet_name=self._sheet_name,
            engine='calamine',
            header=None,
            skiprows=data_start - 1,
            names=self._var_names,
        )

        if self._progress_cb:
            self._progress_cb(90)

        # 后处理：类型转换（对象列→数值，跳过 calamine 已推断的类型）
        datetime_cols = self._detect_datetime_cols(df)
        obj_cols = [
            c for c in df.columns
            if c not in datetime_cols and df[c].dtype == object
        ]
        if obj_cols:
            df[obj_cols] = df[obj_cols].apply(
                pd.to_numeric, errors='coerce'
            )
        for col in datetime_cols:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")

        return df

    def _read_chunks(self) -> pd.DataFrame:
        """优化后的 openpyxl 读取：类型探测 + 预分配数组 + 按 chunk 批量转换。

        内存峰值从 ~3×（all_rows list + chunks list + concat）降到 ~1×（预分配数组）。
        """
        import datetime as _dt
        data_start = self.desc_rows + (3 if self.has_unit else 2)  # 1-based
        max_row = self._ws.max_row or 0
        total_rows = max_row - data_start + 1

        if total_rows <= 0:
            return pd.DataFrame(columns=self._var_names)

        from src.core.config import FLOAT32_REPRESENTABLE_MAX

        # —— 阶段 1：前 100 行做列类型探测
        sample_size = min(100, total_rows)
        sample_iter = self._ws.iter_rows(
            min_row=data_start, max_row=data_start + sample_size - 1,
            values_only=True,
        )
        sample_rows = [next(sample_iter) for _ in range(sample_size)]

        num_cols = len(self._var_names)
        col_dtypes: list[str] = []  # "float32" / "float64" / "datetime" / "object"
        for c in range(num_cols):
            col_sample = [row[c] for row in sample_rows if c < len(row) and row[c] is not None]
            if not col_sample:
                col_dtypes.append("float32")
                continue

            # 检查是否为 datetime 列
            datetime_count = sum(
                1 for val in col_sample if isinstance(val, (_dt.datetime, _dt.date))
            )
            if datetime_count > len(col_sample) * 0.5:
                col_dtypes.append("datetime")
                continue

            # 判断是否为纯数值列
            numeric_count = 0
            max_abs = 0.0
            for val in col_sample:
                try:
                    v = float(val)
                    abs_v = abs(v)
                    if abs_v > max_abs:
                        max_abs = abs_v
                    numeric_count += 1
                except (ValueError, TypeError):
                    break

            if numeric_count == len(col_sample):
                col_dtypes.append(
                    "float32" if max_abs <= FLOAT32_REPRESENTABLE_MAX else "float64"
                )
            else:
                col_dtypes.append("object")

        # —— 阶段 2：预分配 numpy 列数组
        arrays = []
        for dt in col_dtypes:
            if dt.startswith("float"):
                arrays.append(np.full(total_rows, np.nan, dtype=dt))
            elif dt == "datetime":
                arrays.append(np.full(total_rows, np.datetime64("NaT"), dtype="datetime64[ns]"))
            else:
                arrays.append([None] * total_rows)

        # —— 阶段 3：按 chunk 流式读取 + C 级批量转换 + 写入预分配数组
        total_chunks = max(1, (total_rows + self.CHUNK_SIZE - 1) // self.CHUNK_SIZE)
        increment = 80 / total_chunks

        row_iter = self._ws.iter_rows(
            min_row=data_start, max_row=max_row, values_only=True,
        )

        written_rows = 0
        for chunk_idx in range(total_chunks):
            chunk_size = min(self.CHUNK_SIZE, total_rows - written_rows)
            if chunk_size <= 0:
                break

            # 3a: 收集当前 chunk 的行（仅持有 chunk_size 行的 tuple list）
            chunk_rows = []
            for _ in range(chunk_size):
                try:
                    chunk_rows.append(next(row_iter))
                except StopIteration:
                    break
            if not chunk_rows:
                break

            actual_rows = len(chunk_rows)
            start_row = written_rows
            end_row = written_rows + actual_rows

            # 3b: 构建临时 DataFrame（仅 chunk 大小）
            df_chunk = pd.DataFrame(chunk_rows, columns=self._var_names)
            del chunk_rows

            # 3c: 按列类型做批量转换并写入预分配数组
            for c, (col_name, dt) in enumerate(zip(self._var_names, col_dtypes)):
                if c >= len(df_chunk.columns):
                    break
                arr = arrays[c]
                series = df_chunk[col_name]

                if dt.startswith("float"):
                    numeric = pd.to_numeric(series, errors="coerce").to_numpy(dtype=dt, na_value=np.nan)
                    arr[start_row:end_row] = numeric
                elif dt == "datetime":
                    dt_arr = pd.to_datetime(series, errors="coerce").to_numpy(dtype="datetime64[ns]")
                    arr[start_row:end_row] = dt_arr
                else:
                    for i, val in enumerate(series):
                        arr[start_row + i] = val

            del df_chunk
            written_rows = end_row

            if self._progress_cb:
                self._progress_cb(15 + int(min(80, (chunk_idx + 1) * increment)))

        # —— 阶段 4：一次性组装 DataFrame
        df = pd.DataFrame({
            name: arr for name, arr in zip(self._var_names, arrays)
        })

        # —— 阶段 4.5：object 列的数值兜底转换
        obj_cols = [c for c in df.columns if df[c].dtype == object]
        if obj_cols:
            df[obj_cols] = df[obj_cols].apply(pd.to_numeric, errors='coerce')

        # —— 阶段 4.6：复用 _detect_datetime_cols() 做二次确认
        datetime_cols = self._detect_datetime_cols(df)
        for col in datetime_cols:
            if col in df.columns and df[col].dtype == object:
                df[col] = pd.to_datetime(df[col], errors="coerce")

        return df

    @staticmethod
    def _detect_datetime_cols(df_chunk: pd.DataFrame) -> set[str]:
        """从首块数据中检测 datetime 类型列"""
        import datetime as _dt
        dt_cols: set[str] = set()
        for col in df_chunk.columns:
            sample = df_chunk[col].dropna()
            if len(sample) == 0:
                continue
            first_val = sample.iloc[0]
            if isinstance(first_val, (_dt.datetime, _dt.date, _dt.time)):
                dt_cols.add(col)
        return dt_cols

    def _infer_time_columns(self):
        """推断时间列"""
        import datetime as _dt

        time_keywords = ["time", "date", "datetime", "timestamp", "zeit", "tmod"]
        date_candidates = [
            "%H:%M:%S.%f", "%H:%M:%S",
            "%d/%m/%Y", "%Y/%m/%d", "%Y-%m-%d",
        ]

        for col in self._var_names:
            col_lower = col.lower()
            if not any(kw in col_lower for kw in time_keywords):
                continue
            if col not in self._df.columns:
                continue

            s = self._df[col]
            s_sample = s.head(10).dropna()
            if len(s_sample) == 0:
                continue

            # 情况1: 已解析为 datetime 类型
            if pd.api.types.is_datetime64_any_dtype(s):
                self.date_formats[col] = "%Y-%m-%d %H:%M:%S"
                if self.time_column_name is None:
                    self.time_column_name = col
                continue

            # 情况2: 列元素是 Python datetime 对象（object dtype）
            first_val = s_sample.iloc[0]
            if isinstance(first_val, (_dt.datetime, _dt.date)):
                self._df[col] = pd.to_datetime(self._df[col], errors="coerce")
                self.date_formats[col] = "%Y-%m-%d %H:%M:%S"
                if self.time_column_name is None:
                    self.time_column_name = col
                continue

            # 情况3: 仍为字符串，使用格式候选列表匹配
            for fmt in date_candidates:
                try:
                    pd.to_datetime(s_sample, format=fmt, errors="raise")
                    self.date_formats[col] = fmt
                    if self.time_column_name is None:
                        self.time_column_name = col
                    break
                except (ValueError, TypeError):
                    continue

    @staticmethod
    def get_sheet_info(file_path: str) -> list[dict]:
        """获取所有 Sheet 的基本信息（名称、行数、列数）

        使用 openpyxl 只读模式，仅获取元数据，不解析数据内容。
        """
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        result = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result.append({
                'name': sheet_name,
                'rows': ws.max_row or 0,
                'cols': ws.max_column or 0,
            })
        wb.close()
        return result
